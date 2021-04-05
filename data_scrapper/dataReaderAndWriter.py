import time
from data_scrapper.dataFormatter import DataFormatter
from data_scrapper.dataExtractor import Extractor
from data_scrapper import constants
import openpyxl 
import concurrent.futures

class FillDataInExcel:
    # This is Main Function.
    def readAndWriteRequiredCellText(self, excelPath, columnOfRequiredElements, rowFromScrapingStarts, finalCreatedExcelName, totalThreads=10):
        self.startingTime = time.perf_counter()
        
        self.excelPath = excelPath
        self.columnOfRequiredElements = columnOfRequiredElements
        # Column of Required Element is column no. of Resume Id
        self.rowFromScrapingStarts = rowFromScrapingStarts
        # Row from where scraping will start.
        self.finalCreatedExcelName = finalCreatedExcelName
        # rowFromScrapingStarts is the row from we want to scrap data.
        self.totalThreads = totalThreads
        # Total number of threads or chrome windows which has to be opened.
        
        self.excelData = openpyxl.load_workbook(self.excelPath)
        self.requiredExcelSheet = self.excelData[self.excelData.active.title]
        self.totalRows = self.requiredExcelSheet.max_row
        self.totalColumns = self.requiredExcelSheet.max_column
        
        # Adding extra columns to our requiredExcelSheet. 
        # Field to be added : candidateName, email, phoneNumber, status, address, state, country,
        # fields of job or job profile.
        self.valueAtLastColumnOfFirstRow = self.requiredExcelSheet.cell(row=1, column=self.totalColumns).value
        self.valueAtSecondLastColumnOfFirstRow = self.requiredExcelSheet.cell(row=1, column=self.totalColumns - 1).value
        # Below code wil not let to repeat the Constant Row 1 Fields.
        if  self.valueAtLastColumnOfFirstRow != constants.prerequisiteFirstRowTextList[-1] and self.valueAtSecondLastColumnOfFirstRow != constants.prerequisiteFirstRowTextList[-2]: 
            for i in range(1, len(constants.prerequisiteFirstRowTextList) + 1):
                self.requiredExcelSheet.cell(row=1, column=self.totalColumns + i, value=constants.prerequisiteFirstRowTextList[i-1])
        
        # From here we have to make the thread working.
        # Thread Starts From Here. 
        masterListOfIntervalsList = self.rowDividerInEqualPartsAccordingToThreads()
        
        futureList = []
        resultsListOfSucessfulThreads = []
        resultsListOfUnsucessfulThreads = []
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.totalThreads) as threadExecuter:
            for intervalList in masterListOfIntervalsList:
                futures = threadExecuter.submit(self.getAndWriteValuesInExcel, intervalList)
                futureList.append(futures)
            
            for completedFuture in futureList:
                print(completedFuture.result())
                if completedFuture.result() == 'Thread has completed sucessfully.':
                    resultsListOfSucessfulThreads.append(True)
                    
                elif completedFuture.result() == 'Thread has not completed sucessfully.':
                    resultsListOfUnsucessfulThreads.append(False)
                # Threads end here.
            
        # All threads are sucessfully completed.
        if len(resultsListOfSucessfulThreads) == len(futureList):
            self.saveModifiedExcel()
            print(f'All {len(resultsListOfSucessfulThreads)} threads are completed Sucessfully.')
        
        # All threads are not sucessfully completed.
        elif len(resultsListOfUnsucessfulThreads) == len(futureList):
            print(f'All {len(resultsListOfUnsucessfulThreads)} threads are not completed Sucessfully.')
            
        # Some threads are sucessfully completed and some are not.
        elif len(resultsListOfSucessfulThreads) > 0 and len(resultsListOfUnsucessfulThreads) > 0:
            self.saveModifiedExcel()
            print(f'{len(resultsListOfSucessfulThreads)} threads are completed Sucessfully.')
            print(f'{len(resultsListOfUnsucessfulThreads)} threads are not completed Sucessfully.')
            
        
    def rowDividerInEqualPartsAccordingToThreads(self):
        # These are the total number to divide the total row in equal parts.
        # Mainly we have to make some threads for readAndWriteRequiredCellText() method.
        intervalDifference = int((self.totalRows-self.rowFromScrapingStarts)/self.totalThreads)

        listOfTotalIntervalsOfRows = [self.rowFromScrapingStarts]

        # Appending all equal intervals in the list.
        for i in range(1, self.totalThreads):
            listOfTotalIntervalsOfRows.append(self.rowFromScrapingStarts + (i*intervalDifference))
        listOfTotalIntervalsOfRows.append(self.totalRows+1)
        
        # List of starting & ending intervals.
        listOfStartingIntervals = list(filter(lambda x: listOfTotalIntervalsOfRows.index(x)!=len(listOfTotalIntervalsOfRows)-1, listOfTotalIntervalsOfRows))
        listOfEndingIntervals = list(filter(lambda x: listOfTotalIntervalsOfRows.index(x)!=0, listOfTotalIntervalsOfRows))

        # Creating a dictionary of intervals with key=startingInterval & value=endingInterval.
        intervalsDictionary = dict(zip(listOfStartingIntervals, listOfEndingIntervals))
        
        # Appending all the intervals list in master list.
        allIntervalsListInsideMasterList = []
        for key, value in intervalsDictionary.items():
            singleIntervalList = [key, value]
            allIntervalsListInsideMasterList.append(singleIntervalList)
                
        return allIntervalsListInsideMasterList
        
          
    def getAndWriteValuesInExcel(self, singleIntervalsList):
        dataScrapper = Extractor()
        dataScrapper.start(constants.autoLoginWebsiteUrl, isHeadless=True)
        self.singleIntervalsList = singleIntervalsList
        self.isBrowserRunning = dataScrapper.isBrowserRunning
  
        isCompleted = 'Thread has not completed sucessfully.'
        for i in range(self.singleIntervalsList[0], self.singleIntervalsList[1]):
            # Reading the required text(Resume Id) from cell. 
            self.requiredCellText = self.requiredExcelSheet.cell(row=i, column=self.columnOfRequiredElements).value
            
            # Sometimes due to poor excel handling their are many rows are present but all are empty.
            # This will help to scape those empty cells. 
            if self.requiredCellText != None and self.isBrowserRunning:
                # Putting the requried text(Resume Id) and returning the raw Data List. 
                # If raw is not scrapped then it will return None & value of self.rawCandidateDataList will be None.
                self.rawCandidateDataList = dataScrapper.extractRequiredFeildData(self.requiredCellText)
                # This list contains all the data in Raw form of Html & Selenium.
                
                # Is raw data has scrapped or not.
                self.isRawDataScrapped = dataScrapper.isRawDataScrapped
                
                if self.isRawDataScrapped:     
                    try:
                        self.finalDataDictionary = DataFormatter.formatter(self.rawCandidateDataList)
                        # formatter of class DataFormatter will format the Raw data and return a Dictionary(Key & value pair)
                        # of final required data that is Candidate Name, Candidate Email ID, Candidate Phone Number, 
                        # Candidate Status, Candidate Address, Candidate State, Candidate Country & Candidate Field.

                    except Exception as e:
                        print(f'Error : {e}')
                        print(f'Row {i} has not added in excel as raw data has not formatted properly.')
                        continue
                        
                    else:    
                        # Adding the required fields In Excel.
                        # if Prerequisite field are not added then.
                        if self.valueAtLastColumnOfFirstRow != constants.prerequisiteFirstRowTextList[-1] and self.valueAtSecondLastColumnOfFirstRow != constants.prerequisiteFirstRowTextList[-2]: 
                            for x in range(1, len(constants.prerequisiteFirstRowTextList) + 1):
                                self.requiredExcelSheet.cell(row=i, column=self.totalColumns + x, value=self.finalDataDictionary[constants.prerequisiteFirstRowTextList[x-1]])
                            print(f"Sucessfully added Row {i}.")
                            
                        # if Prerequisite field are already added then.
                        if self.valueAtLastColumnOfFirstRow == constants.prerequisiteFirstRowTextList[-1] and self.valueAtSecondLastColumnOfFirstRow == constants.prerequisiteFirstRowTextList[-2]: 
                            for x in range(1, len(constants.prerequisiteFirstRowTextList) + 1):
                                self.requiredExcelSheet.cell(row=i, column=self.totalColumns - len(constants.prerequisiteFirstRowTextList) + x, value=self.finalDataDictionary[constants.prerequisiteFirstRowTextList[x-1]])
                            print(f"Sucessfully added Row {i}.")  
                        
                        # Determine the sucess of thread.
                        isCompleted = 'Thread has completed sucessfully.'
                        
                elif self.isRawDataScrapped == False:     
                    isCompleted = 'Thread has not completed sucessfully.'
                           
            elif self.requiredCellText == None:
                print(f'Row {i} is empty. It has no resume id.')
            else:
                print(f'There is a error while writing the row {i}.')
        
        # Finally shuting down the browser.
        dataScrapper.shutDownBrowser()    
        return isCompleted
        
    def saveModifiedExcel(self):
        # Saving the excel in every entry.
        self.excelData.save(self.finalCreatedExcelName)   
        
        # Total time taken to scrap total rows. 
        self.endingTime = time.perf_counter()
        timeTakenToExecuteProgram = self.endingTime - self.startingTime
        print(f'Total time taken to scrap {self.totalRows - self.rowFromScrapingStarts + 1} Rows is {timeTakenToExecuteProgram : 0.2f} seconds.')
        print(f"The modified Excel is saved with the name of '{self.finalCreatedExcelName}'.")
        